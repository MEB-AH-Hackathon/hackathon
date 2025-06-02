import os
from pathlib import Path
from typing import Optional, Dict, List, Union
import time
import json
import csv

import requests
from dotenv import load_dotenv
from pydantic import BaseModel, ValidationError

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


class AnalyzedText(BaseModel):
    adverse_events: List[str]
    error_type: Optional[str]
    AGE_YRS: Optional[str]
    SEX: Optional[str]
    OTHER_MEDS: Optional[str]
    CUR_ILL: Optional[str]
    HISTORY: Optional[str]
    job: Optional[str]
    lifestyle: Optional[str]
    severity: Optional[str]


class AnthropicCSVAnalyzer:
    API_URL = "https://api.anthropic.com/v1/messages"
    API_VERSION = "2023-06-01"
    DEFAULT_MODEL = "claude-3-haiku-20240307"
    DEFAULT_MAX_TOKENS = 1024

    def __init__(self, api_key: Optional[str] = None):
        load_dotenv()
        self.api_key = api_key or os.getenv("ANTHROPIC_API_KEY")
        if not self.api_key:
            raise ValueError("API key not found. Set ANTHROPIC_API_KEY or pass api_key parameter.")

    def analyze_text(
        self,
        text: str,
        row_data: Dict[str, any],
        prompt: Optional[str] = None,
        model: Optional[str] = None,
        max_tokens: Optional[int] = None
    ) -> Dict[str, any]:
        prompt = prompt or (
            "Extract structured data from the following clinical event description.\n"
            "Return a JSON object with:\n"
            "{\n"
            "  \"adverse_events\": [\"list of symptoms or adverse events\"],\n"
            "  \"error_type\": \"e.g. wrong vaccine, wrong age group, improper storage, etc.\",\n"
            "  \"AGE_YRS\": \"Extracted age if mentioned\",\n"
            "  \"SEX\": \"M/F/unknown\",\n"
            "  \"OTHER_MEDS\": \"any other medications taken\",\n"
            "  \"CUR_ILL\": \"current illnesses or conditions\",\n"
            "  \"HISTORY\": \"medical history if mentioned\",\n"
            "  \"job\": \"occupation or job description\",\n"
            "  \"lifestyle\": \"notable lifestyle habits (e.g. smoking, alcohol)\",\n"
            "  \"severity\": \"severity of the event (mild/moderate/severe/life-threatening)\"\n"
            "}\n\n"
            f"Text: {text}\n\n"
            "IMPORTANT: If any information is not available, not mentioned, or cannot be determined, "
            "you MUST use exactly the string 'unknown' (lowercase) as the value. "
            "Do not use 'none', 'N/A', 'not mentioned', 'not specified', or any other variation. "
            "Use 'unknown' for missing data.\n\n"
            "Respond ONLY with the JSON object."
        )

        payload = {
            "model": model or self.DEFAULT_MODEL,
            "max_tokens": max_tokens or self.DEFAULT_MAX_TOKENS,
            "messages": [{"role": "user", "content": prompt}]
        }

        headers = {
            "x-api-key": self.api_key,
            "anthropic-version": self.API_VERSION,
            "content-type": "application/json"
        }

        response = requests.post(self.API_URL, headers=headers, json=payload)
        response.raise_for_status()
        raw_text = response.json()["content"][0]["text"]

        try:
            parsed_json = json.loads(raw_text)
            structured_data = AnalyzedText(**parsed_json)
            return {
                "row_data": row_data,
                "analyzed_text": text[:100] + "..." if len(text) > 100 else text,
                "response": structured_data.model_dump()
            }
        except (json.JSONDecodeError, ValidationError) as e:
            return {
                "row_data": row_data,
                "analyzed_text": text[:100] + "..." if len(text) > 100 else text,
                "error": str(e),
                "raw_response": raw_text
            }

    def analyze_csv_field(
        self,
        csv_path: Union[str, Path],
        field_name: str,
        delay_between_calls: float = 0.5,
        max_rows: Optional[int] = None
    ) -> Dict[str, List]:
        csv_path = Path(csv_path)
        if not csv_path.exists() or not csv_path.is_file():
            raise ValueError(f"Invalid CSV file path: {csv_path}")

        rows = []
        with open(csv_path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            headers = reader.fieldnames
            if field_name not in headers:
                raise ValueError(f"Field '{field_name}' not found in CSV. Available fields: {headers}")
            for row in reader:
                if row.get(field_name, '').strip():
                    rows.append(row)
                    if max_rows and len(rows) >= max_rows:
                        break

        results = []
        failed = []

        for i, row in enumerate(rows):
            if i > 0:
                time.sleep(delay_between_calls)

            text_to_analyze = row[field_name]
            original_row = row.copy()
            highlight_fields = {}

            try:
                result = self.analyze_text(text_to_analyze, row, None)
                response_data = result.get("response", {})

                # Fields to check and backfill
                fields_to_check = ["AGE_YRS", "SEX", "OTHER_MEDS", "CUR_ILL", "HISTORY"]

                def normalize_for_comparison(value, field_name):
                    """Normalize values for comparison based on field type"""
                    if not value or str(value).strip() == "":
                        return ""
                    
                    value_str = str(value).strip().lower()
                    
                    # Handle SEX field - normalize to M/F format
                    if field_name == "SEX":
                        if value_str in ["male", "m"]:
                            return "m"
                        elif value_str in ["female", "f"]:
                            return "f"
                        elif value_str == "unknown":
                            return "unknown"
                        return value_str
                    
                    # Handle AGE_YRS - treat as float
                    elif field_name == "AGE_YRS":
                        try:
                            # Remove .0 from floats for comparison
                            float_val = float(value_str)
                            if float_val.is_integer():
                                return str(int(float_val))
                            return str(float_val)
                        except ValueError:
                            return value_str
                    
                    return value_str

                for field in fields_to_check:
                    original = row.get(field) or ""
                    original_normalized = normalize_for_comparison(original, field)
                    
                    new_value = response_data.get(field, "unknown") or "unknown"
                    new_value_normalized = normalize_for_comparison(new_value, field)

                    if not original_normalized and new_value_normalized != "unknown":
                        # Field was empty and LLM has actual data, add LLM value
                        if field == "SEX" and new_value_normalized in ["m", "f"]:
                            row[field] = new_value_normalized.upper()  # Store as M/F
                        else:
                            row[field] = new_value
                        highlight_fields[field] = True
                    elif original_normalized != new_value_normalized and new_value_normalized != "unknown":
                        # Values differ and LLM output is not "unknown", add as separate field
                        row[f"{field}_llm_output"] = new_value
                        highlight_fields[f"{field}_llm_output"] = True

                # Add the additional fields only if they have actual data
                for field in ["job", "lifestyle", "severity"]:
                    value = response_data.get(field, "unknown") or "unknown"
                    if value != "unknown":
                        row[field] = value
                        highlight_fields[field] = True

                # Handle adverse events and error type
                adverse_events = response_data.get("adverse_events", [])
                if adverse_events:
                    row["adverse_events"] = "; ".join(adverse_events)
                    highlight_fields["adverse_events"] = True
                
                error_type = response_data.get("error_type", "unknown") or "unknown"
                if error_type != "unknown":
                    row["error_type"] = error_type
                    highlight_fields["error_type"] = True

                result["row_data"] = {
                    "final_row": row,
                    "highlight_fields": highlight_fields
                }
                results.append(result)
                print(f"✓ Processed row {i + 1}")
            except Exception as e:
                failed.append({"row_data": original_row, "error": str(e)})
                print(f"✗ Failed row {i + 1}: {e}")

        return {"successful": results, "failed": failed}

    def export_to_excel(self, results: Dict[str, List], output_file: str = "../data/csv_analysis_results.xlsx"):
        successful_results = results.get("successful", [])
        if not successful_results:
            print("No successful results to export")
            return

        data = []
        highlight_map = []

        for result in successful_results:
            row = result["row_data"]["final_row"]
            highlights = result["row_data"]["highlight_fields"]
            data.append(row)
            highlight_map.append(highlights)

        df = pd.DataFrame(data)
        df.to_excel(output_file, index=False)

        wb = load_workbook(output_file)
        ws = wb.active
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        for row_idx, highlight_fields in enumerate(highlight_map, start=2):  # start=2 to skip header
            for col_idx, key in enumerate(df.columns, start=1):
                if highlight_fields.get(key):
                    ws.cell(row=row_idx, column=col_idx).fill = yellow_fill

        wb.save(output_file)
        print(f"Results exported to Excel: {output_file}")


def main():
    try:
        analyzer = AnthropicCSVAnalyzer()

        results = analyzer.analyze_csv_field(
            csv_path="../data/2025VAERSDATA.csv",
            field_name="SYMPTOM_TEXT",
            max_rows=15,
            delay_between_calls=1.0
        )

        analyzer.export_to_excel(results)

        print("\nSample results:")
        for result in results["successful"][:3]:
            print(json.dumps(result["response"], indent=2))

    except Exception as e:
        print(f"Error: {e}")


if __name__ == "__main__":
    main()
