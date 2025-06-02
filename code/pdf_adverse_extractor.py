import os
from pathlib import Path
from typing import Optional, Dict, List, Union
import time
import json

import requests
from dotenv import load_dotenv
from pydantic import BaseModel, ValidationError
import fitz  # PyMuPDF
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import jsonschema


class AdverseReactionData(BaseModel):
    controlled_trial_text: str
    symptoms_list: List[str]
    study_type: Optional[str]
    source_section: Optional[str]
    full_pdf_text: str


class PDFAdverseExtractor:
    API_URL = "https://api.anthropic.com/v1/messages"
    API_VERSION = "2023-06-01"
    DEFAULT_MODEL = "claude-3-haiku-20240307"
    DEFAULT_MAX_TOKENS = 2048

    def __init__(self, api_key: Optional[str] = None):
        load_dotenv()
        self.api_key = api_key or os.getenv("ANTHROPIC_API_KEY")
        if not self.api_key:
            raise ValueError("API key not found. Set ANTHROPIC_API_KEY or pass api_key parameter.")

    def extract_text_from_pdf(self, pdf_path: Union[str, Path]) -> str:
        """Extract text from PDF using PyMuPDF"""
        pdf_path = Path(pdf_path)
        if not pdf_path.exists():
            raise ValueError(f"PDF file not found: {pdf_path}")
        
        text = ""
        try:
            doc = fitz.open(pdf_path)
            for page in doc:
                text += page.get_text()
            doc.close()
        except Exception as e:
            raise ValueError(f"Error reading PDF {pdf_path}: {e}")
        
        return text

    def extract_adverse_reactions(
        self,
        text: str,
        filename: str,
        model: Optional[str] = None,
        max_tokens: Optional[int] = None
    ) -> Dict[str, any]:
        """Extract adverse reaction data from PDF text using Anthropic API"""
        
        prompt = (
            "Extract adverse reaction information from this pharmaceutical package insert or clinical document.\n\n"
            "IMPORTANT: Only extract information from CONTROLLED CLINICAL TRIALS, SOLICITED ADVERSE REACTIONS, "
            "or other CONTROLLED STUDIES. DO NOT include information from:\n"
            "- VAERS data\n"
            "- Unsolicited adverse events\n"
            "- Post-marketing surveillance\n"
            "- Spontaneous reports\n"
            "- Case reports\n\n"
            "Return a JSON object with:\n"
            "{\n"
            "  \"controlled_trial_text\": \"Direct quote of all text related to adverse reactions from controlled trials/studies\",\n"
            "  \"symptoms_list\": [\"list of specific symptoms/adverse reactions mentioned\"],\n"
            "  \"study_type\": \"type of controlled study (e.g., 'randomized controlled trial', 'clinical trial', 'solicited adverse events')\",\n"
            "  \"source_section\": \"section name where this information was found (e.g., 'Clinical Trials Experience', 'Solicited Adverse Reactions')\"\n"
            "}\n\n"
            "If no controlled trial adverse reaction data is found, return:\n"
            "{\n"
            "  \"controlled_trial_text\": \"unknown\",\n"
            "  \"symptoms_list\": [],\n"
            "  \"study_type\": \"unknown\",\n"
            "  \"source_section\": \"unknown\"\n"
            "}\n\n"
            f"Document text:\n{text}\n\n"  # Send full document text
            "Respond ONLY with the JSON object."
        )

        payload = {
            "model": model or self.DEFAULT_MODEL,
            "max_tokens": max_tokens or self.DEFAULT_MAX_TOKENS,
            "messages": [{"role": "user", "content": prompt}]
        }

        headers = {
            "x-api-key": self.api_key,
            "anthropic-version": self.API_VERSION,
            "content-type": "application/json"
        }

        try:
            response = requests.post(self.API_URL, headers=headers, json=payload)
            response.raise_for_status()
            raw_text = response.json()["content"][0]["text"]

            parsed_json = json.loads(raw_text)
            # Add the full PDF text to the response (not sent to LLM)
            parsed_json["full_pdf_text"] = text
            structured_data = AdverseReactionData(**parsed_json)
            
            return {
                "filename": filename,
                "success": True,
                "data": structured_data.model_dump(),
                "raw_response": raw_text
            }
        except (json.JSONDecodeError, ValidationError, requests.RequestException) as e:
            return {
                "filename": filename,
                "success": False,
                "error": str(e),
                "raw_response": raw_text if 'raw_text' in locals() else None
            }

    def process_pdf_folder(
        self,
        folder_path: Union[str, Path],
        delay_between_calls: float = 1.0,
        max_files: Optional[int] = None
    ) -> Dict[str, List]:
        """Process all PDFs in a folder"""
        folder_path = Path(folder_path)
        if not folder_path.exists() or not folder_path.is_dir():
            raise ValueError(f"Invalid folder path: {folder_path}")

        pdf_files = list(folder_path.glob("*.pdf"))
        if max_files:
            pdf_files = pdf_files[:max_files]

        results = []
        failed = []

        for i, pdf_file in enumerate(pdf_files):
            if i > 0:
                time.sleep(delay_between_calls)

            print(f"Processing {pdf_file.name} ({i+1}/{len(pdf_files)})...")

            try:
                # Extract text from PDF
                text = self.extract_text_from_pdf(pdf_file)
                
                if not text.strip():
                    failed.append({
                        "filename": pdf_file.name,
                        "error": "No text extracted from PDF"
                    })
                    print(f"✗ No text found in {pdf_file.name}")
                    continue

                # Extract adverse reactions
                result = self.extract_adverse_reactions(text, pdf_file.name)
                
                if result["success"]:
                    results.append(result)
                    print(f"✓ Processed {pdf_file.name}")
                else:
                    failed.append(result)
                    print(f"✗ Failed {pdf_file.name}: {result.get('error', 'Unknown error')}")

            except Exception as e:
                failed.append({
                    "filename": pdf_file.name,
                    "error": str(e)
                })
                print(f"✗ Error processing {pdf_file.name}: {e}")

        return {"successful": results, "failed": failed}

    def export_to_excel(self, results: Dict[str, List], output_file: str = "../intermediate_results/pdf_adverse_reactions.xlsx"):
        """Export results to Excel with highlighting"""
        successful_results = results.get("successful", [])
        if not successful_results:
            print("No successful results to export")
            return

        # Prepare data for Excel
        data = []
        for result in successful_results:
            data_row = {
                "filename": result["filename"],
                **result["data"]
            }
            # Convert symptoms list to string
            data_row["symptoms_list"] = "; ".join(data_row["symptoms_list"]) if data_row["symptoms_list"] else ""
            data.append(data_row)

        # Create DataFrame and save to Excel
        df = pd.DataFrame(data)
        df.to_excel(output_file, index=False)

        # Add highlighting for non-unknown values
        wb = load_workbook(output_file)
        ws = wb.active
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        for row_idx in range(2, len(data) + 2):  # Start from row 2 (after header)
            for col_idx in range(1, len(df.columns) + 1):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                # Highlight cells that contain actual data (not "unknown" or empty)
                if cell_value and str(cell_value).strip() and str(cell_value).strip().lower() != "unknown":
                    ws.cell(row=row_idx, column=col_idx).fill = yellow_fill

        wb.save(output_file)
        print(f"Results exported to Excel: {output_file}")

    def validate_results(self, results: Dict[str, List]) -> bool:
        """Validate results against JSON schema"""
        schema_path = "../KEY_INFO/pdf_extraction_schema.json"
        try:
            with open(schema_path, 'r', encoding='utf-8') as f:
                schema = json.load(f)
            
            jsonschema.validate(results, schema)
            print("✓ Results validation passed")
            return True
        except FileNotFoundError:
            print(f"⚠ Schema file not found: {schema_path}")
            return False
        except jsonschema.ValidationError as e:
            print(f"✗ Validation failed: {e.message}")
            print(f"  Path: {' -> '.join(str(p) for p in e.absolute_path)}")
            return False
        except Exception as e:
            print(f"✗ Validation error: {e}")
            return False

    def export_to_json(self, results: Dict[str, List], output_file: str = "../intermediate_results/pdf_extraction_results.json"):
        """Export all results to JSON with validation"""
        # Validate before saving
        if self.validate_results(results):
            with open(output_file, 'w', encoding='utf-8') as f:
                json.dump(results, f, indent=2, ensure_ascii=False)
            print(f"All results saved to JSON: {output_file}")
        else:
            print(f"⚠ Saving results anyway despite validation failure")
            with open(output_file, 'w', encoding='utf-8') as f:
                json.dump(results, f, indent=2, ensure_ascii=False)
            print(f"Results saved to JSON (unvalidated): {output_file}")


def main():
    try:
        extractor = PDFAdverseExtractor()

        # Process PDFs in data folder
        results = extractor.process_pdf_folder(
            folder_path="../data",
            delay_between_calls=1.5,  # Be respectful to API
            max_files=2  # Limit for initial testing
        )

        # Validate and export results
        extractor.export_to_json(results)

        # Print summary
        successful_count = len(results["successful"])
        failed_count = len(results["failed"])
        print(f"\nProcessing complete:")
        print(f"✓ Successful: {successful_count}")
        print(f"✗ Failed: {failed_count}")

        # Show sample results
        if results["successful"]:
            print("\nSample extraction:")
            sample = results["successful"][0]["data"]
            print(f"File: {results['successful'][0]['filename']}")
            print(f"Study type: {sample['study_type']}")
            print(f"Source section: {sample['source_section']}")
            print(f"Symptoms found: {len(sample['symptoms_list'])}")
            if sample['symptoms_list']:
                print(f"Sample symptoms: {', '.join(sample['symptoms_list'][:3])}...")

    except Exception as e:
        print(f"Error: {e}")


if __name__ == "__main__":
    main()